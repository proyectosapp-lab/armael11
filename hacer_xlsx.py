from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F="Arial"
AZUL="1F3864"; AMARILLO="FFF2CC"; GRIS="E8E8E8"; VERDE="E2EFDA"
tit  = Font(name=F,size=11,bold=True,color="FFFFFF")
cab  = PatternFill("solid",fgColor=AZUL)
norm = Font(name=F,size=10)
neg  = Font(name=F,size=10,bold=True)
h1   = Font(name=F,size=14,bold=True,color=AZUL)
gris = Font(name=F,size=10,color="808080")
borde= Border(*[Side(style="thin",color="BFBFBF")]*4)

wb=Wb=Workbook()

# ─────────────────────────── INSTRUCCIONES ───────────────────────────
ins=wb.active; ins.title="Instrucciones"
ins.column_dimensions["A"].width=3
ins.column_dimensions["B"].width=104
def li(r,txt,f=norm):
    ins.cell(r,2,txt).font=f; ins.cell(r,2).alignment=Alignment(wrap_text=True,vertical="top")

li(2,"TSTE · Carga de clubes",h1)
li(4,"Dos hojas: CLUBES (una fila por club) y FUENTES (una fila por fuente de contenido).",neg)
li(5,"Están vinculadas por el campo id: el id que pongas en Clubes es el que va en la columna club_id de Fuentes.")
li(7,"QUÉ COMPLETÁS VOS (celdas en amarillo)",neg)
li(8,"Lo que sabe un hincha y no se puede averiguar solo: colores, patrón de la camiseta, estrellas, "
     "apodos, las trampas de desambiguación, y sobre todo QUÉ FUENTES SEGUIR. Esa curaduría es el trabajo real.")
li(10,"QUÉ NO COMPLETES (celdas en gris)",neg)
li(11,"El channelId de YouTube, la URL exacta del RSS, el id de API-Football y el formato del feed los "
      "resuelvo yo a partir del arroba o del sitio. Si los cargás a mano vas a perder horas y a equivocarte.")
li(13,"CÓMO CARGAR UNA FUENTE",neg)
li(14,"En url_o_handle alcanza con: el arroba de YouTube (@la1913oficial), el link de una playlist, "
      "o el dominio del medio (lavozdelinterior.com.ar). Yo busco el feed y verifico que esté vivo.")
li(16,"EL ORDEN IMPORTA",neg)
li(17,"Cargá primero DOS O TRES clubes completos y avisame. Los corro por el pipeline y confirmamos que el "
      "esquema aguanta antes de que cargues treinta. Si algo falta, es mejor descubrirlo en la fila 3 que en la 30.")
li(19,"Talleres viene cargado entero como ejemplo: mirá esas filas antes de empezar.",neg)

ins.cell(21,2,"Clubes cargados:").font=neg
ins.cell(21,3,"=COUNTA(Clubes!A:A)-1").font=norm
ins.cell(22,2,"Fuentes cargadas:").font=neg
ins.cell(22,3,"=COUNTA(Fuentes!A:A)-1").font=norm
ins.cell(23,2,"Clubes sin ninguna fuente:").font=neg
ins.cell(23,3,'=COUNTIF(Clubes!L2:L200,0)').font=norm
ins.column_dimensions["C"].width=10

# ─────────────────────────── LISTAS ───────────────────────────
lis=wb.create_sheet("Listas")
listas={
 "A":("patron",["bastones","franja","banda","mitades","liso"]),
 "B":("tipo",["oficial","medio","hinchas","periodista","analista"]),
 "C":("plataforma",["youtube","rss","sitio"]),
 "D":("filtro",["ninguno","titulo","texto"]),
 "E":("contexto",["deportes","general"]),
}
for col,(nom,vals) in listas.items():
    lis[f"{col}1"]=nom; lis[f"{col}1"].font=tit; lis[f"{col}1"].fill=cab
    for i,v in enumerate(vals,start=2): lis[f"{col}{i}"]=v; lis[f"{col}{i}"].font=norm
    lis.column_dimensions[col].width=14
lis.sheet_state="hidden"

def encabezar(ws, cols):
    for i,(nom,ancho,color,ayuda) in enumerate(cols,start=1):
        c=ws.cell(1,i,nom); c.font=tit; c.fill=cab; c.border=borde
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=ancho
        a=ws.cell(2,i,ayuda); a.font=gris; a.border=borde
        a.alignment=Alignment(wrap_text=True,vertical="top")
    ws.row_dimensions[1].height=30; ws.row_dimensions[2].height=42
    ws.freeze_panes="A3"

# ─────────────────────────── CLUBES ───────────────────────────
cl=wb.create_sheet("Clubes")
COLS_CL=[
 ("id",16,"am","minúsculas y guiones. ej: talleres-cba"),
 ("nombre",16,"am","como lo dice el hincha"),
 ("nombre_completo",30,"am","el nombre legal, para desambiguar"),
 ("ciudad",16,"am",""),
 ("color_1",11,"am","hex del color principal. ej: #0A3D91"),
 ("color_2",11,"am","hex del segundo color"),
 ("patron",13,"am","cómo es la camiseta"),
 ("estrellas",10,"am","cuántas lleva el escudo. 0 si no lleva"),
 ("apodos",30,"am","separados por coma. ej: la T, el Matador, Albiazul"),
 ("bloqueadores",34,"am","LO MÁS IMPORTANTE: con qué se confunde el nombre. ej: Remedios de Escalada, Perico, taller mecánico"),
 ("apiFootballTeamId",15,"gr","lo completo yo"),
 ("fuentes_cargadas",13,"ve","se calcula solo"),
 ("notas",34,"am",""),
]
encabezar(cl,COLS_CL)
EJEMPLO_CL=["talleres-cba","Talleres","Club Atlético Talleres (Córdoba)","Córdoba",
 "#0A3D91","#FFFFFF","bastones",2,"la T, el Matador, Albiazul, los Tallarines",
 "Remedios de Escalada, Talleres RE, Perico, taller mecánico, talleres municipales, Tigre",
 456,None,"Ejemplo completo. Ojo: 'el Matador' también es Tigre, por eso Tigre va de bloqueador."]
for i,v in enumerate(EJEMPLO_CL,start=1):
    c=cl.cell(3,i,v); c.font=norm; c.border=borde
    c.alignment=Alignment(wrap_text=True,vertical="top")

for r in range(3,201):
    cl.cell(r,12,f"=COUNTIF(Fuentes!A:A,A{r})")
    for i,(nom,_,color,_) in enumerate(COLS_CL,start=1):
        c=cl.cell(r,i)
        c.font=norm; c.border=borde
        c.alignment=Alignment(wrap_text=True,vertical="top")
        if r>3 or c.value is None:
            c.fill=PatternFill("solid",fgColor={"am":AMARILLO,"gr":GRIS,"ve":VERDE}[color])
dv=DataValidation(type="list",formula1="=Listas!$A$2:$A$6",allow_blank=True)
cl.add_data_validation(dv); dv.add(f"G3:G200")

# ─────────────────────────── FUENTES ───────────────────────────
fu=wb.create_sheet("Fuentes")
COLS_FU=[
 ("club_id",16,"am","el mismo id que pusiste en la hoja Clubes"),
 ("nombre",26,"am","cómo se llama la fuente"),
 ("tipo",13,"am","oficial / medio / hinchas / periodista / analista"),
 ("plataforma",12,"am","youtube / rss / sitio"),
 ("url_o_handle",42,"am","el arroba, el link de la playlist, o el dominio del medio. NO hace falta el channelId"),
 ("filtro",12,"am","ninguno = todo lo que sale es del club · titulo = el título siempre lo nombra · texto = hay que desambiguar"),
 ("contexto",12,"am","deportes = solo deporte · general = publica de todo"),
 ("peso",8,"am","0 a 1. Cuánto te importa. 1 = la que leés primero"),
 ("resuelto",30,"gr","lo completo yo: channelId, URL del feed, formato"),
 ("notas",34,"am",""),
]
encabezar(fu,COLS_FU)
EJEMPLO_FU=[
 ["talleres-cba","Canal Showsport · Talleres","medio","rss","canalshowsport.com.ar/futbol/talleres/","ninguno","deportes",1.0,None,"Truco WordPress: cualquier categoría + /feed/ da el RSS de esa sección sola"],
 ["talleres-cba","Talleres · sitio oficial","oficial","rss","clubtalleres.com.ar","ninguno","deportes",0.9,None,"Incluye femenino e institucional"],
 ["talleres-cba","La 1913","hinchas","youtube","@la1913oficial","ninguno","deportes",1.0,None,"Medio partidario acreditado desde 2014"],
 ["talleres-cba","El Diario La T","hinchas","youtube","https://www.youtube.com/playlist?list=PLC55Mqag2NmwZd4wwk1H6v4ER-pp0f-s7","ninguno","deportes",1.0,None,"Programa diario y numerado"],
 ["talleres-cba","Nacho Castellano","periodista","youtube","@nachocastellano_","ninguno","deportes",1.0,None,"Primicias"],
 ["talleres-cba","Pablo Chucrel","analista","youtube","@pablochucrel7","titulo","deportes",0.9,None,"Cubre varios clubes: por eso filtro por título"],
 ["talleres-cba","Mundo Albiazul (La Voz)","medio","youtube","https://www.youtube.com/playlist?list=PLe1HoJfB4goAesjdYrGjVvLDP8Ms7LnfF","ninguno","deportes",1.0,None,"Playlist dedicada dentro de un canal general: sale monotemática gratis"],
 ["talleres-cba","Hoy Día Córdoba · Deportes","medio","rss","hoydia.com.ar/category/deportes/","texto","deportes",0.6,None,"De 20 notas, una es de Talleres. Necesita desambiguación"],
 ["talleres-cba","El Doce","medio","sitio","eldoce.tv","texto","general",0.8,None,"Sin RSS clásico: se lee por su feed de Google News"],
 ["talleres-cba","La Nueva Mañana","medio","sitio","lmdiario.com.ar","texto","general",0.6,None,"Sin RSS: se lee por news sitemap"],
]
for j,fila in enumerate(EJEMPLO_FU):
    for i,v in enumerate(fila,start=1):
        c=fu.cell(3+j,i,v); c.font=norm; c.border=borde
        c.alignment=Alignment(wrap_text=True,vertical="top")
for r in range(3,401):
    for i,(nom,_,color,_) in enumerate(COLS_FU,start=1):
        c=fu.cell(r,i); c.font=norm; c.border=borde
        c.alignment=Alignment(wrap_text=True,vertical="top")
        if c.value is None:
            c.fill=PatternFill("solid",fgColor={"am":AMARILLO,"gr":GRIS}[color])
for col,rango in [("C","=Listas!$B$2:$B$6"),("D","=Listas!$C$2:$C$4"),
                  ("F","=Listas!$D$2:$D$4"),("G","=Listas!$E$2:$E$3")]:
    d=DataValidation(type="list",formula1=rango,allow_blank=True)
    fu.add_data_validation(d); d.add(f"{col}3:{col}400")

wb.save("TSTE-clubes.xlsx")
print("escrito")
